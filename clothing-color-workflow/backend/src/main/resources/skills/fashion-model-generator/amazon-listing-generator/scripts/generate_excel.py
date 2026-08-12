#!/usr/bin/env python3
"""
Amazon Listing + Image Prompt Excel Generator
Called by the amazon-listing-generator skill after producing copy.
Usage: python3 generate_excel.py --product "..." --title "..." ... --output "path.xlsx"
"""

import argparse
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────
def fill(h):
    if len(h) == 6: h = "FF" + h
    return PatternFill("solid", fgColor=h)

def border(c="FFCCCCCC"):
    s = Side(style="thin", color=c)
    return Border(top=s, bottom=s, left=s, right=s)

def wl(i=1):  return Alignment(wrap_text=True, vertical="top",    horizontal="left",   indent=i)
def wc():     return Alignment(wrap_text=True, vertical="center", horizontal="center")
def wlm(i=1): return Alignment(wrap_text=True, vertical="center", horizontal="left",   indent=i)

def hdr(ws, row, text, bg, fg="FFFFFFFF", end="O"):
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Arial", size=11, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = wlm(1)
    ws.row_dimensions[row].height = 26

def label_row(ws, row, cols_labels, bg="FF2C3E7A"):
    for ci, (col, label, width) in enumerate(cols_labels):
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        c.fill = fill(bg)
        c.alignment = wc()
        c.border = border()
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 36

def data_cell(ws, row, col, val, bg="FFFFFFFF", fg="222222", bold=False, sz=9, height=None):
    c = ws[f"{col}{row}"]
    c.value = val
    c.font = Font(name="Arial", size=sz, bold=bold, color=f"FF{fg}")
    c.fill = fill(bg)
    c.alignment = wl(1)
    c.border = border()
    if height:
        ws.row_dimensions[row].height = height
    return c

# ════════════════════════════════════════════════════════════════════════════
IMAGE_SLOTS = [
    ("AM-01","主图 Main Image",      "3000x3000px","Real photo required"),
    ("AS-02","核心卖点图",            "2000x2000px","AI allowed"),
    ("AS-03","功能拆解图",            "2000x2000px","AI allowed"),
    ("AS-04","尺寸参数图",            "2000x2000px","AI allowed"),
    ("AS-05","场景使用图",            "2000x2000px","AI allowed"),
    ("AS-06","细节特写图 (2x2)",      "2000x2000px","AI allowed"),
    ("AS-07","竞品对比图",            "2000x2000px","AI allowed"),
    ("AS-08","安装/使用步骤图",       "2000x2000px","AI allowed"),
    ("AS-09","包装全家福",            "2000x2000px","AI allowed"),
    ("AV-01","主图视频脚本",          "1920x1080px","Video script"),
]


def normalize_prompts(raw):
    if isinstance(raw, dict):
        return [raw.get(slot, "") for slot, *_ in IMAGE_SLOTS]
    if isinstance(raw, list):
        prompts = list(raw[:len(IMAGE_SLOTS)])
        prompts.extend([""] * (len(IMAGE_SLOTS) - len(prompts)))
        return prompts
    raise ValueError("--prompts-json must contain a list or a dict keyed by slot id")


def build_workbook(args):
    wb = Workbook()

    # ── SHEET 1: LISTING ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Listing"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF6900"

    ws.merge_cells("A1:O1")
    ws["A1"] = f"Amazon Listing  ·  {args.product}  ·  AI Generated"
    ws["A1"].font  = Font(name="Arial", size=13, bold=True, color="FFFFFFFF")
    ws["A1"].fill  = fill("1A1A2E")
    ws["A1"].alignment = wc()
    ws.row_dimensions[1].height = 36

    # TITLE
    hdr(ws, 2, "▌ TITLE", "FFBF360C")
    label_row(ws, 3,
        [("A","字段",14),("B","内容",70),("C","字符数",10),("D","合规检查",28)],
        bg="FFE64A19")
    data_cell(ws, 4, "A", "Title", bg="FFE8EAF6", bold=True)
    data_cell(ws, 4, "B", args.title, bg="FFECF4FF", fg="003399", height=32)
    ws[f"C4"].value = f"=LEN(B4)"
    ws[f"C4"].font = Font(name="Arial", size=9, color="FF880000")
    ws[f"C4"].fill = fill("FFFAFAFA"); ws[f"C4"].alignment = wc(); ws[f"C4"].border = border()
    data_cell(ws, 4, "D",
        "□ ≤150 chars  □ Primary keyword in first 80  □ No ALL CAPS  □ No promo words",
        bg="FFF1F8E9", fg="1B5E20", sz=8)

    # BULLETS
    hdr(ws, 6, "▌ BULLET POINTS  (5×)", "FF1565C0")
    label_row(ws, 7,
        [("A","#",5),("B","Bullet 内容",70),("C","字符",8),("D","COSMO维度",22),("E","合规",20)],
        bg="FF1976D2")
    bullets = [args.b1, args.b2, args.b3, args.b4, args.b5]
    cosmo_tags = [
        "capableOf + causes",
        "hasProperty + distinguishedFrom",
        "suitableFor + usedInContext",
        "motivatedBy + distinguishedFrom",
        "partOf + relatedTo",
    ]
    checks = [
        "□ ALL-CAPS label  □ 150-200 chars  □ Benefit first",
        "□ Material spec included  □ Competitor contrast",
        "□ Target user named  □ Scene described",
        "□ Category concern addressed  □ No competitor names",
        "□ All items listed  □ Guarantee/warranty included",
    ]
    for i, (b, cosmo, chk) in enumerate(zip(bullets, cosmo_tags, checks)):
        r = 8 + i
        bg = "FFF5F5F5" if i % 2 == 0 else "FFFFFFFF"
        data_cell(ws, r, "A", f"B{i+1}", bg="FFE3F2FD", bold=True, fg="0D47A1")
        data_cell(ws, r, "B", b, bg="FFECF4FF", fg="003399", height=52)
        ws[f"C{r}"].value = f"=LEN(B{r})"
        ws[f"C{r}"].font = Font(name="Arial",size=9,color="FF880000")
        ws[f"C{r}"].fill = fill("FFFAFAFA"); ws[f"C{r}"].alignment = wc(); ws[f"C{r}"].border = border()
        data_cell(ws, r, "D", cosmo, bg="FFE8EAF6", fg="1A237E", sz=8)
        data_cell(ws, r, "E", chk, bg="FFF1F8E9", fg="1B5E20", sz=8)

    # DESCRIPTION
    hdr(ws, 14, "▌ DESCRIPTION  (target 1500-2000 chars)", "FF4A148C")
    label_row(ws, 15,
        [("A","内容",100),("B","字符数",10),("C","合规检查",28)],
        bg="FF6A1B9A")
    data_cell(ws, 16, "A", args.description, bg="FFECF4FF", fg="003399", height=120)
    ws["B16"].value = "=LEN(A16)"
    ws["B16"].font = Font(name="Arial",size=9,color="FF880000")
    ws["B16"].fill = fill("FFFAFAFA"); ws["B16"].alignment = wc(); ws["B16"].border = border()
    data_cell(ws, 16, "C",
        "□ ≥1500 chars  □ 5 paragraphs  □ Pain→Solution→Features→Scenes→Guarantee  □ Natural language",
        bg="FFF1F8E9", fg="1B5E20", sz=8)

    # BACKEND
    hdr(ws, 18, "▌ BACKEND SEARCH TERMS  (≤250 bytes, space-separated)", "FF1B5E20")
    label_row(ws, 19,
        [("A","Backend Terms",100),("B","字节数",10),("C","合规检查",28)],
        bg="FF388E3C")
    data_cell(ws, 20, "A", args.backend, bg="FFECF4FF", fg="003399", height=40)
    ws["B20"].value = len(args.backend.encode("utf-8"))
    ws["B20"].font = Font(name="Arial",size=9,color="FF880000")
    ws["B20"].fill = fill("FFFAFAFA"); ws["B20"].alignment = wc(); ws["B20"].border = border()
    data_cell(ws, 20, "C",
        "□ ≤250 bytes  □ Space-separated only  □ No repeats from Title/Bullets  □ No brand names",
        bg="FFF1F8E9", fg="1B5E20", sz=8)

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "A3"

    # ── SHEET 2: IMAGE PROMPTS ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Image Prompts")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "2196F3"

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Image Prompt Kit  ·  {args.product}  ·  9 Slots + Video"
    ws2["A1"].font  = Font(name="Arial", size=13, bold=True, color="FFFFFFFF")
    ws2["A1"].fill  = fill("1A237E")
    ws2["A1"].alignment = wc()
    ws2.row_dimensions[1].height = 36

    # Column headers
    img_cols = [("A","#",5),("B","图片名称",18),("C","尺寸",13),
                ("D","AI可用",9),("E","提示词 (复制→粘贴到AI工具)",70),("F","状态",9)]
    for col, label, width in img_cols:
        c = ws2[f"{col}2"]
        c.value = label
        c.font  = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        c.fill  = fill("2C3E7A")
        c.alignment = wc()
        c.border = border()
        ws2.column_dimensions[col].width = width
    ws2.row_dimensions[2].height = 28

    prompts_list = getattr(args, "prompts_list", [""] * len(IMAGE_SLOTS))

    for i, ((slot, name, size, ai_flag), prompt) in enumerate(zip(IMAGE_SLOTS, prompts_list)):
        r = 3 + i
        bg = "FFF5F5F5" if i % 2 == 0 else "FFFFFFFF"
        ai_bg  = "FFFFEBEE" if "Real photo" in ai_flag else ("FFFFF8E1" if "Video" in ai_flag else "FFF1F8E9")
        ai_fg  = "CC0000"   if "Real photo" in ai_flag else ("E65100" if "Video" in ai_flag else "1B5E20")

        data_cell(ws2, r, "A", slot,     bg="FFE3F2FD", bold=True, fg="0D47A1")
        data_cell(ws2, r, "B", name,     bg=bg)
        data_cell(ws2, r, "C", size,     bg=bg, sz=8)
        c = ws2[f"D{r}"]; c.value = ai_flag
        c.font = Font(name="Arial", size=8, bold=True, color=f"FF{ai_fg}")
        c.fill = fill(ai_bg); c.alignment = wc(); c.border = border()
        data_cell(ws2, r, "E", prompt,   bg="FFECF4FF" if prompt else bg, fg="003399", height=70)
        c2 = ws2[f"F{r}"]; c2.value = "待生成" if "Real photo" not in ai_flag else "待拍摄"
        c2.font = Font(name="Arial", size=8, bold=True, color="FF5D4037")
        c2.fill = fill("FFFFF9C4"); c2.alignment = wc(); c2.border = border()

    # ── SHEET 3: CHECKLIST ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("发布检查清单")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "43A047"

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Amazon 发布前检查清单  ·  Listing + 图片"
    ws3["A1"].font  = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
    ws3["A1"].fill  = fill("2E7D32")
    ws3["A1"].alignment = wc()
    ws3.row_dimensions[1].height = 32

    checklist = [
        ("LISTING", [
            ("Title", "≤150字符，主关键词在前80字符，无违禁词，自然可读"),
            ("Title", "2025年新规：无全大写、无特殊符号、无促销词(Best/Free/Sale)"),
            ("Bullets", "5条，每条150-200字符，全大写标签开头，利益优先"),
            ("Bullets", "前400字符合计覆盖最重要信息（手机端截断点）"),
            ("Description", "1500-2000字符，5段式结构，语言自然，无关键词堆砌"),
            ("Backend", "≤250字节，空格分隔，无重复词，无品牌名"),
            ("COSMO", "≥8个COSMO语义维度已覆盖（见算法说明Sheet）"),
            ("Rufus", "Listing可回答'谁用/在哪用/解决什么问题'"),
        ]),
        ("IMAGES", [
            ("主图AM-01", "真实拍摄，纯白背景#FFFFFF，产品≥85%，无文字/logo，≥1600px"),
            ("主图AM-01", "150px缩略图测试：手机搜索结果可识别产品"),
            ("副图×8", "每张图内容与Listing对应文案一致（数字/材质/功能）"),
            ("副图×8", "AI生成图可用于AS-02至AS-09，不可用于主图"),
            ("所有图", "推荐3000×3000px，最低1600×1600px，JPEG RGB格式"),
            ("A+", "A+图片不能与主图/副图重复（否则审核被拒）"),
            ("视频", "≥720p，≤5GB，MP4/MOV，45-90秒，前5秒有产品亮相"),
        ]),
        ("上架", [
            ("Seller Central", "分类节点准确（影响COSMO语义分类）"),
            ("属性字段", "Target Audience / Intended Use / Subject Matter 填写完整"),
            ("价格", "参考BSR Top20竞品定价区间，首次上架可低5-10%获评价"),
            ("FBA/FBM", "FBA优先（影响Buy Box获取）"),
            ("上线后", "第1周检查Title是否被Amazon自动修改（2025年新规）"),
            ("监控", "每2周检查关键词排名，每季度重新做竞品ASIN分析"),
        ]),
    ]

    for col, w in [("A", 12), ("B", 22), ("C", 60)]:
        ws3.column_dimensions[col].width = w

    r = 2
    for section, items in checklist:
        ws3.merge_cells(f"A{r}:C{r}")
        c = ws3[f"A{r}"]
        c.value = f"  ▌ {section}"
        c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill  = fill("388E3C" if section == "IMAGES" else ("1565C0" if section == "LISTING" else "FF6F00"))
        c.alignment = wlm(1); c.border = border()
        ws3.row_dimensions[r].height = 22; r += 1

        for field, desc in items:
            c = ws3[f"A{r}"]; c.value = "□"
            c.font = Font(name="Arial", size=11); c.fill = fill("FFFFFFFF")
            c.alignment = wc(); c.border = border()
            c = ws3[f"B{r}"]; c.value = field
            c.font = Font(name="Arial", size=8, bold=True, color="FF222222")
            c.fill = fill("FFF5F5F5"); c.alignment = wl(1); c.border = border()
            c = ws3[f"C{r}"]; c.value = desc
            c.font = Font(name="Arial", size=8, color="FF444444")
            c.fill = fill("FFFFFFFF"); c.alignment = wl(1); c.border = border()
            ws3.row_dimensions[r].height = 20; r += 1
        r += 1  # spacer

    # ── Sheet order ───────────────────────────────────────────────────────────
    order = ["Listing", "Image Prompts", "发布检查清单"]
    smap  = {s.title: s for s in wb._sheets}
    wb._sheets = [smap[n] for n in order if n in smap]

    return wb


# ════════════════════════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--product",     default="Product")
    p.add_argument("--title",       default="")
    p.add_argument("--b1",          default="")
    p.add_argument("--b2",          default="")
    p.add_argument("--b3",          default="")
    p.add_argument("--b4",          default="")
    p.add_argument("--b5",          default="")
    p.add_argument("--description", default="")
    p.add_argument("--backend",     default="")
    p.add_argument("--output",      default="Amazon_Listing.xlsx")
    # Image prompts as a JSON file path (optional — skill writes this)
    p.add_argument("--prompts-json", default=None)
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    # Load image prompts if provided
    if args.prompts_json:
        with open(args.prompts_json, encoding="utf-8") as f:
            args.prompts_list = normalize_prompts(json.load(f))
    else:
        args.prompts_list = ["(Prompt will be filled in by Codex)"] * len(IMAGE_SLOTS)

    wb = build_workbook(args)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True) if output.parent != Path(".") else None
    wb.save(output)
    print(f"✅  Saved: {args.output}")
